import openpyxl as px
import numpy as np
import matplotlib.pyplot as plt
#import plotly.plotly as py
import math

Male=[]
Female=[]

def calculateProbability(x,mean,stdev):
    exponent=math.exp(-(math.pow(x-mean,2)/(2*math.pow(stdev,2))))
    return (1/(math.sqrt(2*math.pi) * stdev)) * exponent

def Bayesian(x,menmean,menstd,femalemean,femalestd):
    for k in [55,60,65,70,75,80]:
      num1=(len(Female)*calculateProbability(k,femalemean,femalestd))
      num2=(len(Male)*calculateProbability(k,menmean,menstd))
      print(num1/(num1+num2))

def Build1DHistogramClassifier(X,T,B,xmin,xmax):
   HF=np.zeros(B).astype('int32');
   HM=np.zeros(B).astype('int32');
   bindices=[(np.round(((B-1)*(x-xmin)/(xmax-xmin)))).astype('int32') for x in X];
   for i,b in enumerate(bindices):
       if T[i]=='Female':
          HF[b]+=1;
          Female.append(X[i]);
       else:
          HM[b]+=1;
          Male.append(X[i]);

  # histogram=plt.figure()
  # plt.hist(Male,32,alpha=0.5)
  # plt.hist(Female,32,alpha=0.5)
  # plt.show()


   SSmen=sum(HM)
   SSwomen=sum(HF)
   print(SSmen)
   print(SSwomen)
  
   for k in [55,60,65,70,75,80]: 
     bin=(np.round(((B-1)*(k-xmin)/(xmax-xmin)))).astype('int32')
     print(HF[bin],HM[bin])
     prob=(HF[bin]/(HM[bin] + HF[bin])).astype('float')
  
     print "prob is %d" % k + ", %d" % bin + ", %0.2f" % prob
  

   print(HM)
   print(HF) 

   return[HF,HM]


def Build1DBayesianClassifier(X,T,xmin,xmax):
    
    femaleMean=np.mean(Female,dtype=np.float64)
    maleMean=np.mean(Male,dtype=np.float64)
    print(maleMean)
    print(femaleMean)
    femaleSTD=np.std(Female,dtype=np.float64)
    maleSTD=np.std(Male,dtype=np.float64)
    print(maleSTD)
    print(femaleSTD)
    
    print("Bayesian")
    Bayesian(55,maleMean,maleSTD,femaleMean,femaleSTD)


W =px.load_workbook('1.xlsx')
p = W.get_sheet_by_name(name='Data')

i=0
row_count = p.max_row
column_count = p.max_column

#print row_count
#print column_count

X=[]
T=[]



for i in range(0,50): 
    X.append(int(p.cell(row=i+2,column =1).value)*12+int(p.cell(row=i+2,column=2).value))
    T.append(p.cell(row=i+2,column=3).value)

min_height=52
max_height=83

print(min_height)
print(max_height)


Build1DHistogramClassifier(X,T,32,min_height,max_height);
Build1DBayesianClassifier(X,T,min_height,max_height)
